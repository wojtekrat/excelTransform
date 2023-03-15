import os
import shutil
from openpyxl import load_workbook
from tkinter import *
from tkinter import filedialog as fd
from tkinter.ttk import Progressbar
import datetime

root = Tk()
root.title("Excel manager")
root.resizable(False, False)

main_frame = Frame(root)
main_frame.pack(padx=50, pady=50)

frame_excel = Frame(main_frame, highlightbackground="grey", highlightthickness=1)
frame_excel.pack(padx=5, pady=5)
db_button = Button(frame_excel, text="Choose Excel file", command=lambda: open_file())
db_button.grid(row=1, column=0, padx=5, pady=5)
db_name_label1 = Label(frame_excel, text='Excel file not chosen')
db_name_label1.grid(row=2, column=0, padx=5, pady=5)

frame_excel2 = Frame(main_frame, highlightbackground="grey", highlightthickness=1)
frame_excel2.pack(padx=5, pady=5)
db_name_label = Label(frame_excel2, text="Choose sheet: ")
db_name_label.grid(row=1, column=0, padx=5, pady=5)
sheet_variable = StringVar()
sheet_dropdown = OptionMenu(frame_excel2, sheet_variable, [])
sheet_dropdown.grid(row=2, column=0, padx=5, pady=5)

frame_folder1 = Frame(main_frame, highlightbackground="grey", highlightthickness=1)
frame_folder1.pack(padx=5, pady=5)
open_button = Button(frame_folder1, text="Choose folder with files", command=lambda: open_folder())
open_button.grid(row=1, column=0, padx=5, pady=5)
db_name_label2 = Label(frame_folder1, text='Folder with files has not been chosen')
db_name_label2.grid(row=2, column=0, padx=5, pady=5)

frame_folder2 = Frame(main_frame, highlightbackground="grey", highlightthickness=1)
frame_folder2.pack(padx=5, pady=5)
open_button2 = Button(frame_folder2, text="Choose output folder", command=lambda: select_output_folder())
open_button2.grid(row=1, column=0, padx=5, pady=5)
db_name_label3 = Label(frame_folder2, text='Output folder has not been chosen')
db_name_label3.grid(row=2, column=0, padx=5, pady=5)

open_button3 = Button(main_frame, height= 5, width=20, text="Copy files", command=lambda: transform_excel())
open_button3.pack(padx=5, pady=5)

frame_progress = Frame(main_frame, highlightbackground="grey", highlightthickness=1 )
frame_progress.pack(padx=5, pady=5)
progress_label = Label(frame_progress, text="File copy progress")
progress_label.grid(row=1, column=0, padx=5, pady=5)
ramka = Frame(frame_progress)
ramka.grid(row=2, column=0)
progress_bar = Progressbar(ramka, orient=HORIZONTAL, length=100, mode='determinate', maximum=100)
progress_bar.pack(pady=10)

error_label = Label(main_frame, text='')
error_label.pack(padx=5, pady=5)


def open_file():
    global xlsx_path, sheet_names, choice

    # Select file and update path label
    filetype = [('Excel files', '*.xlsx;*.xlsm;*.xltx;*.xltm')]
    filepath = fd.askopenfilenames(filetypes=filetype)
    filepath2 = ''
    for i in filepath:
        filepath2 = filepath2 + i
        xlsx_path = filepath2
    db_name_label1.config(text=xlsx_path)

    # Load all sheet names from the Excel file
    wb = load_workbook(xlsx_path)
    sheet_names = wb.sheetnames

    # Create a dropdown menu to choose a sheet
    choice = StringVar()
    sheet_dropdown = OptionMenu(frame_excel2, choice, *sheet_names)
    sheet_dropdown.grid(row=2, column=0, padx=5, pady=5)

    def update_choice(*args):
        global choice
        choice = choice.get()
    choice.trace('w', update_choice)

    choice.set(sheet_names[0])  # Set default value


def open_folder():
    global folder_path
    mr_dir = fd.askdirectory()
    folder_path = mr_dir
    db_name_label2.config(text=folder_path)

def select_output_folder():
    global output_folder 
    output_folder = fd.askdirectory()
    db_name_label3.config(text=output_folder)
    return output_folder


def transform_excel():
    global xlsx_path, folder_path, choice, output_folder
    now = datetime.datetime.now()
    # Check if the XLSX file path is valid
    if not os.path.isfile(xlsx_path) or not xlsx_path.endswith('.xlsx'):
        with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
            f.write(f'Invalid Excel file\n {now}')
        error_label.config(text='Invalid Excel file')
        return
        
    # Check if the folder path is valid
    if not os.path.isdir(folder_path):
        with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
            f.write(f'Invalid files path\n {now}')
        error_label.config(text='Invalid files path')
        return
            
    workbook = load_workbook(xlsx_path)
    worksheet_name = choice
    
    try:
        worksheet = workbook[worksheet_name]
    except KeyError:
        with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
            f.write(f'Invalid sheet name: {worksheet_name} {now}\n')
        error_label.config(text=f'Invalid sheet name: {worksheet_name}')
        return

    # Loop through the rows of the worksheet
    for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        progress_bar['value'] = (i / (worksheet.max_row - 1)) * 100
        progress_bar.update()
        # Get the values from the columns
        try:
            col1, col2, col3, col4 = row
        except ValueError:
            with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
                f.write(f'Invalid data in row: {row}\n {now}')
            error_label.config(text=f'Invalid data in row: {row}')
            continue

        # Determine the child folder path
        if col2 and col3 and col4:
            child_folder = os.path.join(output_folder, col3  + '_' + now.strftime('%Y-%m-%d'), str(col4))
        else:
            child_folder = os.path.join(output_folder, 'no parameter' + '_' + now.strftime('%Y-%m-%d'))
        os.makedirs(child_folder, exist_ok=True)

        # Loop through the files in the specific folder and copy them to the child folder
        specific_folder = os.path.join(os.getcwd(), folder_path)
        for root, dirs, files in os.walk(specific_folder):
            for filename in files:
                if col1 in filename:
                    src_path = os.path.join(root, filename)
                    if col2:
                        dst_filename = str(col1) + '_' + str(col2) + os.path.splitext(filename)[1]
                    else:
                        dst_filename = col1 + os.path.splitext(filename)[1]
                    dst_path = os.path.join(child_folder, dst_filename)
                    if os.path.exists(dst_path):
                        with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
                            f.write(f'File {dst_path} already exists\n {now}\n')
                        error_label.config(text=f'File {dst_path} already exists')
                        continue
                    try:
                        shutil.copy(src_path, dst_path)
                        error_label.config(text='Files copied properly')
                    except Exception as e:
                        with open(os.path.join(output_folder, 'errors.txt'), 'a+') as f:
                            f.write(f'file copy error {src_path} to {dst_path}: {e} {now}\n')
                        error_label.config(text=f'File copy error {src_path} to {dst_path}: {e}')


root.mainloop()
